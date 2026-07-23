#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_a_datajs.py — Convierte un Excel de operativos de playas en el archivo
playas/data.js que consume el dashboard (window.DEFAULT_DATA).

Replica EXACTAMENTE la lógica de lectura del navegador (playas/index.html:
funciones prettyPeriod y cleanSheet), para que el resultado publicado sea
idéntico al que verías subiendo el archivo en la web.

USO BÁSICO
----------
  # Añade / actualiza las temporadas del Excel dentro del data.js existente:
  python3 scripts/xlsx_a_datajs.py datos_nuevos.xlsx

  # Varios archivos a la vez (p. ej. todos los de la carpeta datos/):
  python3 scripts/xlsx_a_datajs.py datos/*.xlsx

  # Genera un data.js SOLO con lo del Excel (reemplaza todo, como hace la web):
  python3 scripts/xlsx_a_datajs.py datos_nuevos.xlsx --reemplazar

  # Previsualizar sin escribir / sin crear respaldo (CI):
  python3 scripts/xlsx_a_datajs.py datos.xlsx --dry-run
  python3 scripts/xlsx_a_datajs.py datos/*.xlsx --sin-respaldo

REGLAS DEL EXCEL (iguales que en la web)
----------------------------------------
- Cada HOJA = una temporada. El nombre de la hoja debe ser:
      semana.santa26 · semanasanta26 · verano24 · invierno2021 ...
  (empieza por 'semana santa' / 'verano' / 'invierno' + año de 2 o 4 dígitos).
- Columnas por NOMBRE (en cualquier orden): Estado, Destino, Playa, Sitio,
  Clasificación. El encabezado se busca en las primeras 5 filas.
- Estado/Destino/Playa se "arrastran" hacia abajo si la celda está vacía.
- 'Sitio' es obligatorio por fila; sin sitio, la fila se ignora.
- Clasificación: texto que contenga 'no apta' -> 0, 'apta' -> 1, otro/vacío -> null.
"""

import argparse
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# ------------------------------------------------------------------ utilidades
def norm(s):
    """minúsculas + sin acentos + trim  (equiv. al norm() del navegador)."""
    s = "" if s is None else str(s)
    s = s.strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s


def g(v):
    """valor de celda como texto recortado ('' si None)."""
    return "" if v is None else str(v).strip()


def pretty_period(name):
    """Nombre de hoja -> etiqueta legible. Equivale a prettyPeriod() del navegador."""
    m = re.match(
        r"^(semana\.?\s*santa|verano|invierno)\s*[._-]?\s*(\d{2}|\d{4})$",
        str(name).strip().lower(),
    )
    if m:
        temp = re.sub(r"\s+", " ", m.group(1).replace(".", " ")).strip()
        label = "Semana Santa" if temp == "semana santa" else temp[:1].upper() + temp[1:]
        yr = "20" + m.group(2) if len(m.group(2)) == 2 else m.group(2)
        return f"{label} {yr}"
    # respaldo genérico
    out = re.sub(r"[._]+", " ", str(name))
    return re.sub(r"\b\w", lambda c: c.group(0).upper(), out)


def clean_sheet(ws):
    """Hoja de openpyxl -> lista de filas {est,dst,playa,sitio,apta}.
    Equivale a cleanSheet() del navegador."""
    aoa = [list(r) for r in ws.iter_rows(values_only=True)]
    if not aoa:
        return []

    header_row, col = -1, {}
    for i in range(min(len(aoa), 5)):
        r = aoa[i] or []
        found = {}
        for ci, v in enumerate(r):
            n = norm(v)
            if n == "estado":
                found["est"] = ci
            elif n.startswith("destino"):
                found["dst"] = ci
            elif n == "playa":
                found["playa"] = ci
            elif n.startswith("sitio") and "sitio" not in found:
                found["sitio"] = ci
            elif n.startswith("clasificacion") and "clas" not in found:
                found["clas"] = ci
        if "sitio" in found and "clas" in found:
            header_row, col = i, found
            break

    if header_row < 0:
        return []  # formato no reconocido

    est = dst = playa = None
    rows = []
    for i in range(header_row + 1, len(aoa)):
        r = aoa[i]
        if r is None:
            continue

        def cell(key):
            idx = col.get(key)
            return r[idx] if (idx is not None and idx < len(r)) else None

        if "est" in col and g(cell("est")):
            est = g(cell("est"))
        if "dst" in col and g(cell("dst")):
            dst = g(cell("dst"))
        if "playa" in col and g(cell("playa")):
            playa = g(cell("playa"))

        sitio = g(cell("sitio"))
        if not sitio:
            continue
        s = g(cell("clas")).lower()
        apta = 0 if "no apta" in s else (1 if "apta" in s else None)
        rows.append({"est": est, "dst": dst, "playa": playa, "sitio": sitio, "apta": apta})
    return rows


# ------------------------------------------------------------------ data.js I/O
def cargar_data_js(path):
    """Lee playas/data.js y devuelve el dict (o {'periods': []} si no existe)."""
    if not path.exists():
        return {"periods": []}
    txt = path.read_text(encoding="utf-8")
    txt = txt.split("=", 1)[1].strip()          # quita 'window.DEFAULT_DATA ='
    txt = txt.strip().rstrip(";").strip()        # quita ';' final
    return json.loads(txt)


def escribir_data_js(path, data):
    """Escribe data.js con el mismo formato que el original (una sola línea)."""
    payload = json.dumps(data, ensure_ascii=False, separators=(", ", ": "))
    path.write_text(f"window.DEFAULT_DATA = {payload};", encoding="utf-8")


# ------------------------------------------------------------------ principal
def main(argv=None):
    ap = argparse.ArgumentParser(description="Convierte un Excel de playas en playas/data.js")
    ap.add_argument("xlsx", nargs="+", help="Uno o más archivos .xlsx (cada hoja = una temporada)")
    ap.add_argument(
        "--data",
        default=str(Path(__file__).resolve().parent.parent / "playas" / "data.js"),
        help="Ruta a data.js (por defecto: playas/data.js del repo)",
    )
    ap.add_argument(
        "--reemplazar",
        action="store_true",
        help="Genera data.js SOLO con las temporadas de los Excel (descarta las existentes)",
    )
    ap.add_argument("--dry-run", action="store_true", help="No escribe; solo muestra el resumen")
    ap.add_argument("--sin-respaldo", action="store_true",
                    help="No crea el archivo data.js.bak (útil en CI / GitHub Actions)")
    args = ap.parse_args(argv)

    xlsx_paths = [Path(x) for x in args.xlsx]
    faltantes = [str(p) for p in xlsx_paths if not p.exists()]
    if faltantes:
        sys.exit("ERROR: no existe(n) el/los archivo(s): " + ", ".join(faltantes))
    data_path = Path(args.data)

    # 1) Leer los Excel -> periods nuevos (en orden; hojas repetidas ganan las últimas)
    nuevos, orden = {}, []
    for xp in sorted(xlsx_paths, key=lambda p: p.name):
        wb = load_workbook(xp, read_only=True, data_only=True)
        for name in wb.sheetnames:
            rows = clean_sheet(wb[name])
            if not rows:
                print(f"  ⚠ [{xp.name}] hoja '{name}' ignorada (sin datos válidos / formato no reconocido)")
                continue
            if name not in nuevos:
                orden.append(name)
            nuevos[name] = {"key": name, "label": pretty_period(name), "rows": rows}
        wb.close()

    nuevos = [nuevos[k] for k in orden]
    if not nuevos:
        sys.exit("ERROR: no se encontraron hojas con datos válidos en los Excel.")

    print(f"\nTemporadas leídas ({len(nuevos)}):")
    for p in nuevos:
        aptas = sum(1 for r in p["rows"] if r["apta"] == 1)
        noaptas = sum(1 for r in p["rows"] if r["apta"] == 0)
        sindato = sum(1 for r in p["rows"] if r["apta"] is None)
        print(f"  • {p['key']:<18} -> {p['label']:<20} {len(p['rows']):>4} sitios "
              f"(aptas {aptas}, no aptas {noaptas}, sin dato {sindato})")

    # 2) Combinar con el data.js existente
    if args.reemplazar:
        data = {"periods": nuevos}
        print("\nModo --reemplazar: el data.js contendrá SOLO lo del Excel.")
    else:
        data = cargar_data_js(data_path)
        existentes = {p["key"]: i for i, p in enumerate(data["periods"])}
        actualizadas, agregadas = [], []
        for p in nuevos:
            if p["key"] in existentes:
                data["periods"][existentes[p["key"]]] = p
                actualizadas.append(p["key"])
            else:
                data["periods"].append(p)
                agregadas.append(p["key"])
        print(f"\nMezcla con {data_path.name}: "
              f"{len(agregadas)} agregada(s), {len(actualizadas)} actualizada(s).")
        if agregadas:
            print("  agregadas:  " + ", ".join(agregadas))
        if actualizadas:
            print("  actualizadas: " + ", ".join(actualizadas))

    print(f"\nTotal de temporadas en el resultado: {len(data['periods'])}")

    # 3) Escribir
    if args.dry_run:
        print("\n--dry-run: no se escribió nada.")
        return

    if data_path.exists() and not args.reemplazar and not args.sin_respaldo:
        bak = data_path.with_suffix(
            data_path.suffix + ".bak." + datetime.now().strftime("%Y%m%d_%H%M%S"))
        bak.write_text(data_path.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"\nRespaldo del data.js anterior: {bak.name}")

    escribir_data_js(data_path, data)
    print(f"✔ Escrito: {data_path}")
    print("\nSiguiente paso (publicar):")
    print("  git add playas/data.js && git commit -m \"Actualiza datos de playas\" && git push")


if __name__ == "__main__":
    main()
