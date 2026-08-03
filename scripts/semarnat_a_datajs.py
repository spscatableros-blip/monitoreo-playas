#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
semarnat_a_datajs.py — Convierte el reporte "Monitoreo Playas SEMARNAT" (un Excel con
UNA HOJA POR AÑO, formato ancho con celdas combinadas) en playas/data.js
(window.DEFAULT_DATA).

Este Excel es el MAESTRO completo (todos los años/temporadas), así que este script
REGENERA el data.js entero (reemplaza, no mezcla).

USO
---
  # Requisito: pip install openpyxl
  python3 scripts/semarnat_a_datajs.py "Monitoreo Playas SEMARNAT 2013-2026.xlsx"
  python3 scripts/semarnat_a_datajs.py archivo.xlsx --dry-run
  python3 scripts/semarnat_a_datajs.py archivo.xlsx --incluir-2013   # incluir 2013 (preliminar)

FORMATO ESPERADO
----------------
- Una hoja por año (nombre de hoja = año de 4 dígitos: 2014, 2024, ...).
- En cada hoja, una fila de encabezado con: No., Estado, Destino turístico, Playa,
  Sitio de muestreo, Coordenadas, y luego un BLOQUE de 3 columnas por temporada
  (Fecha de muestreo · NMP/100 mL · Clasificación). El nombre de la temporada va en
  el encabezado del bloque: "Monitoreo prevacacional {Semana Santa|Verano|Invierno} {año}".
- Estado / Destino / Playa vienen combinadas (se arrastran hacia abajo).
- Clasificación: texto con "apta" -> 1, "no apta" -> 0, vacío/otro -> null.

Por defecto se OMITE 2013 (el reporte lo marca como preliminar). Usa --incluir-2013 para incluirlo.
"""

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

TEMP_KEY = {"semana santa": "semana.santa", "verano": "verano", "invierno": "invierno"}
TEMP_LABEL = {"semana.santa": "Semana Santa", "verano": "Verano", "invierno": "Invierno"}
TEMP_RANK = {"semana.santa": 0, "verano": 1, "invierno": 2}


def _norm(s):
    s = "" if s is None else str(s)
    s = " ".join(s.strip().lower().split())
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _g(row, i):
    v = row[i] if (i is not None and i < len(row)) else None
    return "" if v is None else str(v).strip()


def _num(v):
    """Valor de NMP/100 mL -> número o None.
    '<10' -> 10 · '>24196' -> 24196 · 'No se midió' / 'N/A' / fecha / vacío -> None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("<", "").replace(">", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _buscar_col(rows, hr, bc, fin, pred):
    """Busca en las sub-filas de encabezado (hr+1, hr+2) la primera columna del rango
    [bc, fin) cuyo texto normalizado cumpla `pred`. Devuelve su índice o None."""
    for off in (1, 2):
        ri = hr + off
        if ri >= len(rows):
            continue
        r = rows[ri]
        for ci in range(bc, min(fin, len(r))):
            if pred(_norm(r[ci])):
                return ci
    return None


def parse_hoja_anio(ws):
    """Devuelve {key: [rows]} para una hoja-año del reporte SEMARNAT."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # 1) localizar la fila de encabezado (contiene 'estado' y una col que empieza con 'sitio')
    hr = -1
    for i in range(min(10, len(rows))):
        cells = [_norm(c) for c in rows[i]]
        if "estado" in cells and any(c.startswith("sitio") for c in cells):
            hr = i
            break
    if hr < 0:
        return {}

    head = [_norm(c) for c in rows[hr]]
    col = {}
    for ci, c in enumerate(head):
        if c == "estado":
            col["est"] = ci
        elif c.startswith("destino"):
            col["dst"] = ci
        elif c == "playa":
            col["playa"] = ci
        elif c.startswith("sitio"):
            col["sitio"] = ci

    # 2) bloques de temporada: buscar "... (Semana Santa|Verano|Invierno) AÑO" en el encabezado
    bloques = []
    for ci, c in enumerate(rows[hr]):
        m = re.search(r"(semana santa|verano|invierno)\s*(\d{4})", _norm(c))
        if m:
            bloques.append((ci, TEMP_KEY[m.group(1)], m.group(2)))
    if not bloques:
        return {}

    # 3) columnas de NMP/100 mL y Clasificación de cada bloque. El reporte tiene dos filas de
    #    sub-encabezado (hr+1: "Calidad bacteriológica…"; hr+2: "Fecha · NMP/100 mL · Clasificación"),
    #    así que se buscan en ambas. Si no aparecen, se usa la posición estándar del bloque.
    ancho = max((len(r) for r in rows[hr:hr + 3]), default=len(head))
    periodos = {}
    for bi, (bc, temp, yr) in enumerate(bloques):
        fin = bloques[bi + 1][0] if bi + 1 < len(bloques) else max(ancho, bc + 3)
        clas = _buscar_col(rows, hr, bc, fin, lambda n: n.startswith("clasificacion"))
        nmp = _buscar_col(rows, hr, bc, fin, lambda n: "nmp" in n)
        periodos[f"{temp}{yr[2:]}"] = {
            "clas": bc + 2 if clas is None else clas,
            "nmp":  bc + 1 if nmp is None else nmp,
            "temp": temp, "yr": yr, "rows": []}

    # 4) filas de datos (con arrastre de Estado/Destino/Playa)
    est = dst = playa = None
    for r in rows[hr + 2:]:
        if _g(r, col.get("est")):
            est = _g(r, col.get("est"))
        if _g(r, col.get("dst")):
            dst = _g(r, col.get("dst"))
        if _g(r, col.get("playa")):
            playa = _g(r, col.get("playa"))
        sitio = _g(r, col.get("sitio"))
        if not sitio:
            continue
        for info in periodos.values():
            s = _g(r, info["clas"]).lower()
            apta = 0 if "no apta" in s else (1 if "apta" in s else None)
            nmp = _num(r[info["nmp"]] if info["nmp"] < len(r) else None)
            info["rows"].append({"est": est, "dst": dst, "playa": playa, "sitio": sitio,
                                 "apta": apta, "nmp": nmp})

    return periodos


def construir(xlsx_path, incluir_2013):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    todos = {}   # key -> info
    for name in wb.sheetnames:
        if not re.fullmatch(r"\d{4}", str(name).strip()):
            continue
        if str(name).strip() == "2013" and not incluir_2013:
            continue
        for key, info in parse_hoja_anio(wb[name]).items():
            if info["rows"]:
                todos[key] = info
    wb.close()

    if not todos:
        sys.exit("ERROR: no se encontraron hojas-año con datos válidos (¿es el reporte SEMARNAT?).")

    # ordenar cronológicamente: por año y por temporada (SS < Verano < Invierno)
    orden = sorted(todos.values(), key=lambda i: (int(i["yr"]), TEMP_RANK[i["temp"]]))
    periods = [{"key": f"{i['temp']}{i['yr'][2:]}",
                "label": f"{TEMP_LABEL[i['temp']]} {i['yr']}",
                "rows": i["rows"]} for i in orden]
    return {"periods": periods}


def main(argv=None):
    ap = argparse.ArgumentParser(description="Convierte el reporte SEMARNAT en playas/data.js")
    ap.add_argument("xlsx", help="Excel del reporte (una hoja por año)")
    ap.add_argument(
        "--data",
        default=str(Path(__file__).resolve().parent.parent / "playas" / "data.js"),
        help="Ruta a playas/data.js (por defecto: la del repo)",
    )
    ap.add_argument("--incluir-2013", action="store_true", help="Incluir el año 2013 (preliminar)")
    ap.add_argument("--dry-run", action="store_true", help="No escribe; solo muestra el resumen")
    args = ap.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"ERROR: no existe el archivo {xlsx_path}")

    data = construir(xlsx_path, args.incluir_2013)

    print(f"\nTemporadas generadas ({len(data['periods'])}):")
    for p in data["periods"]:
        a = sum(1 for r in p["rows"] if r["apta"] == 1)
        n = sum(1 for r in p["rows"] if r["apta"] == 0)
        s = sum(1 for r in p["rows"] if r["apta"] is None)
        print(f"  • {p['key']:<16} {p['label']:<20} {len(p['rows']):>4} sitios "
              f"(aptas {a}, no aptas {n}, sin dato {s})")

    if args.dry_run:
        print("\n--dry-run: no se escribió nada.")
        return

    data_path = Path(args.data)
    payload = json.dumps(data, ensure_ascii=False, separators=(", ", ": "))
    data_path.write_text(f"window.DEFAULT_DATA = {payload};", encoding="utf-8")
    print(f"\n✔ Escrito: {data_path}")
    print("\nSiguiente paso (publicar):")
    print("  git add playas/data.js && git commit -m \"Actualiza datos de playas\" && git push")


if __name__ == "__main__":
    main()
