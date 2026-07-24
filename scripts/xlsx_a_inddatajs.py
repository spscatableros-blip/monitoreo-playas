#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_a_inddatajs.py — Convierte el Excel de la base de calidad de sitios costeros
(CONAGUA) en el archivo indicadores/data.js que consume el tablero de indicadores
(window.IND_DATA).

A diferencia de playas, aquí el Excel es un VOLCADO COMPLETO de la base (todas las
muestras 2012–…), así que este script REGENERA el data.js entero cada vez (reemplaza,
no mezcla). El formato de salida es columnar (arreglos paralelos), reproducido
bit a bit del data.js original.

USO
---
  # Requisito: pip install openpyxl
  python3 scripts/xlsx_a_inddatajs.py indicadores-cost.xlsx
  python3 scripts/xlsx_a_inddatajs.py indicadores-cost.xlsx --dry-run

FORMATO DEL EXCEL
-----------------
Una hoja llamada 'indicadores' (o la única/primera hoja del libro) con encabezados
en la primera fila. Se usan estas columnas, localizadas POR NOMBRE (no por posición):
  - CLAVE SITIO        -> para contar sitios distintos (meta.sitios)
  - FECHA REALIZACIÓN  -> de aquí sale el mes (1–12)
  - Año                -> año de la muestra
  - ESTADO             -> nombre del estado (se recorta; debe ser costero, ver CATALOGO)
  - Las 8 columnas de indicadores: ENTEROC_FEC, SST, OD_%_SUP, OD_mg/L_SUP,
    OD_%_MED, OD_mg/L_MED, OD_%_FON, OD_mg/L_FON

REGLAS (idénticas a como se generó el data.js original)
-------------------------------------------------------
- Las filas cuyo ESTADO (ya recortado) NO está en el catálogo de 17 estados costeros
  se DESCARTAN (p. ej. 'SAN LUIS POTOSÍ' o celdas vacías). Esto también limpia
  variantes con espacios sobrantes ('COLIMA ' -> 'COLIMA').
- Valores de indicadores: los textos de límite de detección se numerizan quitando el
  signo: '<3' -> 3, '>24196' -> 24196, '<1' -> 1. Vacío / no numérico -> null.
- El mes se toma del mes de FECHA REALIZACIÓN.
"""

import argparse
import datetime
import json
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


# 17 estados costeros (orden y grafía EXACTOS del data.js original)
CATALOGO = [
    "BAJA CALIFORNIA", "BAJA CALIFORNIA SUR", "CAMPECHE", "CHIAPAS", "COLIMA",
    "GUERRERO", "JALISCO", "MICHOACÁN DE OCAMPO", "NAYARIT", "OAXACA",
    "QUINTANA ROO", "SINALOA", "SONORA", "TABASCO", "TAMAULIPAS",
    "VERACRUZ DE IGNACIO DE LA LLAVE", "YUCATÁN",
]

# Los 8 indicadores en orden fijo (así los espera el tablero)
INDICADORES = [
    "ENTEROC_FEC", "SST", "OD_%_SUP", "OD_mg/L_SUP",
    "OD_%_MED", "OD_mg/L_MED", "OD_%_FON", "OD_mg/L_FON",
]

META_DEFAULT = {"umbral_enteroc": 200, "fuente": "CONAGUA · BD calidad costeros"}


def _norm(s):
    """MAYÚSCULAS, sin acentos, sin espacios extra: para comparar encabezados/estados."""
    s = "" if s is None else str(s)
    s = " ".join(s.strip().upper().split())
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def numerizar(x):
    """Valor de indicador -> float o None. '<3'->3, '>24196'->24196, ''/texto->None."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("<", "").replace(">", "").replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def elegir_hoja(wb):
    """Devuelve la hoja 'indicadores' (por nombre normalizado) o la única/primera."""
    for name in wb.sheetnames:
        if _norm(name) == "INDICADORES":
            return wb[name]
    return wb[wb.sheetnames[0]]


def localizar_columnas(header):
    """Mapa nombre_lógico -> índice de columna, buscando por encabezado."""
    col = {}
    ind_norm = {_norm(k): k for k in INDICADORES}
    for ci, h in enumerate(header):
        n = _norm(h)
        if n == "CLAVE SITIO":
            col["sitio"] = ci
        elif n.startswith("FECHA REALIZACION"):
            col["fecha"] = ci
        elif n in ("ANO", "ANIO"):
            col["anio"] = ci
        elif n == "ESTADO":
            col["estado"] = ci
        elif n in ind_norm and ind_norm[n] not in col:
            col[ind_norm[n]] = ci
    return col


def cargar_meta_previa(data_path):
    """Conserva umbral_enteroc y fuente del data.js actual si existe."""
    meta = dict(META_DEFAULT)
    if data_path.exists():
        try:
            txt = data_path.read_text(encoding="utf-8")
            d = json.loads(txt.split("=", 1)[1].strip().rstrip(";").strip())
            for k in ("umbral_enteroc", "fuente"):
                if k in d.get("meta", {}):
                    meta[k] = d["meta"][k]
        except Exception:
            pass
    return meta


def construir(xlsx_path, meta_extra):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = elegir_hoja(wb)
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        sys.exit("ERROR: la hoja está vacía.")
    col = localizar_columnas(header)

    faltan = [c for c in ("fecha", "anio", "estado") if c not in col] \
        + [k for k in INDICADORES if k not in col]
    if faltan:
        sys.exit("ERROR: no se encontraron estas columnas en el Excel: " + ", ".join(faltan))

    idx = {_norm(s): i for i, s in enumerate(CATALOGO)}
    est, anio, mes = [], [], []
    v = {k: [] for k in INDICADORES}
    sitios = set()
    descartados = {}
    total = 0

    for r in it:
        total += 1
        e = _norm(r[col["estado"]])
        if e not in idx:
            etq = "(vacío)" if not str(r[col["estado"]] or "").strip() else str(r[col["estado"]]).strip()
            descartados[etq] = descartados.get(etq, 0) + 1
            continue
        est.append(idx[e])
        anio.append(r[col["anio"]])
        f = r[col["fecha"]]
        mes.append(f.month if isinstance(f, datetime.datetime) else None)
        if "sitio" in col:
            sitios.add(r[col["sitio"]])
        for k in INDICADORES:
            v[k].append(numerizar(r[col[k]]))

    wb.close()

    if not est:
        sys.exit("ERROR: no quedó ninguna muestra válida (¿nombres de estado fuera del catálogo?).")

    anios_validos = [a for a in anio if isinstance(a, int)]
    meta = {
        "muestras": len(est),
        "sitios": len(sitios),
        "anio_min": min(anios_validos),
        "anio_max": max(anios_validos),
        **meta_extra,
    }
    data = {"estados": CATALOGO, "est": est, "anio": anio, "mes": mes,
            "indicadores": INDICADORES, "v": v, "meta": meta}
    return data, total, descartados


def main(argv=None):
    ap = argparse.ArgumentParser(description="Convierte el Excel de calidad costera en indicadores/data.js")
    ap.add_argument("xlsx", help="Excel de la base (hoja 'indicadores')")
    ap.add_argument(
        "--data",
        default=str(Path(__file__).resolve().parent.parent / "indicadores" / "data.js"),
        help="Ruta a indicadores/data.js (por defecto: la del repo)",
    )
    ap.add_argument("--dry-run", action="store_true", help="No escribe; solo muestra el resumen")
    args = ap.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"ERROR: no existe el archivo {xlsx_path}")
    data_path = Path(args.data)

    data, total, descartados = construir(xlsx_path, cargar_meta_previa(data_path))

    print(f"\nFilas leídas: {total}")
    print(f"Muestras válidas: {data['meta']['muestras']}  |  Sitios: {data['meta']['sitios']}  "
          f"|  Años: {data['meta']['anio_min']}–{data['meta']['anio_max']}")
    if descartados:
        print("Filas descartadas (estado fuera del catálogo de 17 costeros):")
        for k, n in sorted(descartados.items(), key=lambda x: -x[1]):
            print(f"  • {k}: {n}")

    if args.dry_run:
        print("\n--dry-run: no se escribió nada.")
        return

    payload = json.dumps(data, ensure_ascii=False, separators=(", ", ": "))
    data_path.write_text(f"window.IND_DATA = {payload};", encoding="utf-8")
    print(f"\n✔ Escrito: {data_path}")
    print("\nSiguiente paso (publicar):")
    print("  git add indicadores/data.js && git commit -m \"Actualiza indicadores\" && git push")


if __name__ == "__main__":
    main()
